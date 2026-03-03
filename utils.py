# -*- coding: utf-8 -*-
"""Utilidades compartidas entre páginas: estilos CSS, constantes de color e iconos."""

from __future__ import annotations

from dataclasses import asdict
from io import BytesIO
from typing import Any, Dict

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Colores e iconos por tecnología (paleta Apple-inspired)
# ─────────────────────────────────────────────────────────────────────────────

TECH_COLORS = {
    'Diesel': '#FF9500',
    'Eléctrico - Carga nocturna': '#34C759',
    'Eléctrico - Carga flash': '#007AFF',
    'Eléctrico - Carga por oportunidad': '#5856D6',
    'Hidrógeno': '#AF52DE',
}

TECH_ICONS = {
    'Diesel': '⛽',
    'Eléctrico - Carga nocturna': '🔋',
    'Eléctrico - Carga flash': '⚡',
    'Eléctrico - Carga por oportunidad': '🔌',
    'Hidrógeno': '💧',
}

# Orden canónico de las tecnologías
TECH_ORDER = ['diesel', 'overnight', 'flash', 'opportunity', 'hydrogen']

TECH_NAMES = {
    'diesel': 'Diesel',
    'overnight': 'Eléctrico - Carga nocturna',
    'flash': 'Eléctrico - Carga flash',
    'opportunity': 'Eléctrico - Carga por oportunidad',
    'hydrogen': 'Hidrógeno',
}

TECH_SHORT = {
    'Diesel': 'Diesel',
    'Eléctrico - Carga nocturna': 'Nocturna',
    'Eléctrico - Carga flash': 'Flash',
    'Eléctrico - Carga por oportunidad': 'Oportunidad',
    'Hidrógeno': 'Hidrógeno',
}


# ─────────────────────────────────────────────────────────────────────────────
# CSS global estilo Apple
# ─────────────────────────────────────────────────────────────────────────────

GLOBAL_CSS = """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    }
    
    .stApp {
        background: linear-gradient(180deg, #fafafa 0%, #f5f5f7 100%);
    }
    
    h1 {
        font-weight: 600 !important;
        letter-spacing: -0.5px !important;
        color: #1d1d1f !important;
    }
    
    h2, h3 {
        font-weight: 500 !important;
        color: #1d1d1f !important;
    }
    
    [data-testid="stSidebar"] {
        background: #ffffff !important;
        border-right: 1px solid #e5e5e7 !important;
    }
    
    [data-testid="stMetricValue"] {
        font-size: 28px !important;
        font-weight: 600 !important;
        color: #1d1d1f !important;
    }
    
    [data-testid="stMetricLabel"] {
        font-size: 12px !important;
        font-weight: 500 !important;
        color: #86868b !important;
        text-transform: uppercase !important;
        letter-spacing: 0.5px !important;
    }
    
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: transparent;
    }
    
    .stTabs [data-baseweb="tab"] {
        background: #f5f5f7 !important;
        border-radius: 8px !important;
        padding: 8px 16px !important;
        font-weight: 500 !important;
        border: none !important;
    }
    
    .stTabs [aria-selected="true"] {
        background: #1d1d1f !important;
        color: white !important;
    }
    
    .metric-card {
        background: white;
        border-radius: 16px;
        padding: 24px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04), 0 4px 12px rgba(0,0,0,0.04);
        border: 1px solid #e5e5e7;
        margin-bottom: 16px;
        text-align: center;
    }
    
    hr {
        border: none !important;
        height: 1px !important;
        background: #e5e5e7 !important;
        margin: 32px 0 !important;
    }
    
    .stDataFrame {
        border-radius: 12px !important;
        overflow: hidden !important;
    }
</style>
"""


def inject_css():
    """Inyecta los estilos CSS globales en la página."""
    st.markdown(GLOBAL_CSS, unsafe_allow_html=True)


def format_eur(value: float) -> str:
    """Formatea un valor como euros con separador de miles."""
    if abs(value) >= 1_000_000:
        return f"€{value:,.0f}"
    return f"€{value:,.0f}"


def plotly_layout_defaults() -> dict:
    """Retorna los defaults de layout Plotly consistentes con el estilo Apple."""
    return dict(
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        xaxis=dict(showgrid=False, showline=False),
        yaxis=dict(showgrid=True, gridcolor='#e5e5e7', showline=False),
        margin=dict(t=60, b=40, l=40, r=20),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
    )


# ─────────────────────────────────────────────────────────────────────────────
# Exportación Excel
# ─────────────────────────────────────────────────────────────────────────────

# Colores de tecnología sin '#' para openpyxl
_TECH_HEX = {k: v.lstrip('#') for k, v in TECH_COLORS.items()}

# Etiquetas legibles para métricas operativas
_OP_LABELS: dict[str, str] = {
    'flota_headway': 'Flota mínima (headway)',
    'flota_km': 'Flota mínima (autonomía)',
    'flota_requerida': 'Flota requerida',
    'km_comerciales_dia': 'Km comerciales / día',
    'km_totales_dia': 'Km totales / día',
    'km_por_bus_com': 'Km comerciales / bus / día',
    'km_por_bus_tot': 'Km totales / bus / día',
    'ciclos_por_bus': 'Ciclos / bus / día',
    'autonomia_km': 'Autonomía (km)',
    'autonomia_total_km': 'Autonomía total (km)',
    'autonomia_usable_km': 'Autonomía usable (km)',
    'autonomia_efectiva_km': 'Autonomía efectiva con mini-cargas (km)',
    'km_faltantes': 'Km faltantes (sin mini-carga)',
    'km_recuperados_por_carga_cabecera': 'Km recup. / carga cabecera',
    'km_recuperados_por_carga_trazado': 'Km recup. / carga trazado',
    'mini_cargas_por_bus': 'Mini-cargas / bus / día',
    'cargas_cab_por_bus': 'Cargas cabecera / bus',
    'cargas_traz_por_bus': 'Cargas trazado / bus',
    'cargas_por_ciclo': 'Cargas por ciclo',
    'n_puntos_cabecera': 'Puntos de carga cabecera',
    'n_puntos_trazado': 'Puntos de carga trazado',
    'tiempo_regulacion_min': 'Tiempo regulación (min)',
    'tiempo_carga_cabecera_min': 'Tiempo carga cabecera (min)',
    'tiempo_carga_trazado_min': 'Tiempo carga trazado (min)',
    'n_puntos_carga': 'Puntos de carga totales',
    'cargadores_por_punto_cabecera': 'Carg. por punto cabecera',
    'cargadores_por_punto_trazado': 'Carg. por punto trazado',
    'n_cargadores_cabecera': 'Cargadores cabecera (total)',
    'n_cargadores_trazado': 'Cargadores trazado (total)',
    'n_cargadores_ruta': 'Cargadores en ruta (total)',
    'cargadores_patio': 'Cargadores en patio',
    'energia_consumida_por_bus_kwh': 'Energía consumida / bus (kWh)',
    'energia_recarga_por_bus_kwh': 'Energía recarga / bus (kWh)',
    'energia_mini_cargas_por_bus_kwh': 'Energía mini-cargas / bus (kWh)',
    'energia_total_patio_kwh': 'Energía total patio (kWh/día)',
    'energia_total_ruta_kwh': 'Energía total ruta (kWh/día)',
    'energia_total_dia_kwh': 'Energía total día (kWh)',
    'tiempo_carga_por_bus_h': 'Tiempo carga / bus (h)',
    'potencia_promedio_patio_kw': 'Potencia promedio patio (kW)',
    'potencia_instalada_patio_kw': 'Potencia instalada patio (kW)',
    'potencia_instalada_cabecera_kw': 'Potencia instalada cabecera (kW)',
    'potencia_instalada_trazado_kw': 'Potencia instalada trazado (kW)',
    'potencia_total_instalada_kw': 'Potencia total instalada (kW)',
    'combustible_total_l': 'Combustible total (L/día)',
    'combustible_por_bus_l': 'Combustible / bus (L/día)',
    'h2_total_kg': 'H₂ total (kg/día)',
    'h2_por_bus_kg': 'H₂ / bus (kg/día)',
}

# Etiquetas legibles para inputs de cada dataclass
_INPUT_LABELS: dict[str, str] = {
    # GeneralInputs
    'km_trazado_sentido': 'Longitud por sentido (km)',
    'velocidad_kmh': 'Velocidad comercial (km/h)',
    'headway_min': 'Headway (min)',
    'tiempo_servicio_min': 'Tiempo de servicio (min)',
    'tiempo_entre_servicios_min': 'Regulación en terminal (min)',
    'km_vacio_frac': 'Km en vacío (fracción)',
    # Diesel
    'consumo_litros_km': 'Consumo (L/km)',
    'autonomia_km': 'Autonomía (km)',
    # Eléctricos
    'bateria_kwh': 'Batería (kWh)',
    'consumo_kwh_km': 'Consumo (kWh/km)',
    'soc_reserva_frac': 'SOC reserva (fracción)',
    'cargador_kw': 'Cargador (kW)',
    'eficiencia_carga': 'Eficiencia de carga',
    'ventana_carga_h': 'Ventana de carga (h)',
    'cargador_ruta_kw': 'Cargador ruta (kW)',
    'tiempo_carga_cabecera_min': 'Tiempo carga cabecera (min)',
    'tiempo_carga_trazado_min': 'Tiempo carga trazado (min)',
    'tiempo_regulacion_min': 'Tiempo regulación (min)',
    'cargador_patio_kw': 'Cargador patio (kW)',
    'max_mini_cargas_restriccion': 'Máx. mini-cargas (restricción)',
    # Hidrógeno
    'consumo_h2_kg_km': 'Consumo H₂ (kg/km)',
    # Costos generales
    'horizonte_anios': 'Horizonte (años)',
    'dias_operacion_anio': 'Días operación / año',
    # CAPEX
    'vehiculo_eur': 'Vehículo (€)',
    'infraestructura_deposito_eur': 'Infra depósito (€)',
    'cargador_pistola_eur': 'Cargador pistola (€/ud)',
    'subestacion_electrica_eur': 'Subestación eléctrica (€)',
    'cargador_pantografo_cabecera_eur': 'Cargador pantógrafo cabecera (€/ud)',
    'cargador_pistola_patio_eur': 'Cargador pistola patio (€/ud)',
    'cargador_oportunidad_cabecera_eur': 'Cargador oportunidad cabecera (€/ud)',
    'estacion_hidrogeno_eur': 'Estación hidrógeno (€)',
    # OPEX
    'combustible_eur_litro': 'Diésel (€/litro)',
    'energia_eur_kwh': 'Energía (€/kWh)',
    'mantenimiento_eur_km': 'Mantenimiento (€/km)',
    'bateria_mantenimiento_anual_eur_bus': 'Batería mant. anual (€/bus)',
    'bateria_reemplazo_eur': 'Reemplazo batería (€)',
    'bateria_vida_util_anios': 'Vida útil batería (años)',
    'hidrogeno_eur_kg': 'Hidrógeno (€/kg)',
}

# Etiquetas legibles para métricas de costos
_COST_LABELS: dict[str, str] = {
    'vehiculos': 'Vehículos',
    'cargadores_cabecera': 'Cargadores cabecera',
    'cargadores_patio': 'Cargadores patio',
    'subestacion': 'Subestación eléctrica',
    'estacion_h2': 'Estación hidrógeno',
    'infra_deposito': 'Infraestructura depósito',
    'total_capex': 'TOTAL CAPEX',
    'combustible_energia': 'Combustible / Energía',
    'mantenimiento': 'Mantenimiento',
    'bateria_anual': 'Batería (mant. anual)',
    'total_opex_anual': 'TOTAL OPEX anual',
}

# Estilos reutilizables
_THIN_BORDER = Border(
    left=Side(style='thin', color='D5D5D7'),
    right=Side(style='thin', color='D5D5D7'),
    top=Side(style='thin', color='D5D5D7'),
    bottom=Side(style='thin', color='D5D5D7'),
)
_SECTION_FILL = PatternFill(start_color='F0F0F2', end_color='F0F0F2', fill_type='solid')
_SECTION_FONT = Font(name='Calibri', bold=True, size=11)
_HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
_NORMAL_FONT = Font(name='Calibri', size=11)
_LABEL_FONT = Font(name='Calibri', size=11, color='555555')
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
_TOTAL_FILL = PatternFill(start_color='E8E8EA', end_color='E8E8EA', fill_type='solid')
_TOTAL_FONT = Font(name='Calibri', bold=True, size=11, color='1D1D1F')


def _auto_width(ws) -> None:
    """Ajusta el ancho de columnas al contenido."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ''
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def _write_section_header(ws, row: int, text: str, n_cols: int) -> int:
    """Escribe una fila de encabezado de sección (ej. 'CAPEX')."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _SECTION_FILL
        cell.font = _SECTION_FONT
        cell.border = _THIN_BORDER
        cell.alignment = _LEFT
    ws.cell(row=row, column=1, value=text)
    return row + 1


def _format_value(v: Any) -> Any:
    """Redondea floats a 2 decimales para presentación."""
    if isinstance(v, float):
        return round(v, 2)
    return v


def _tech_header_fill(tech_name: str) -> PatternFill:
    """Retorna el fill de color para el encabezado de una tecnología."""
    hex_color = _TECH_HEX.get(tech_name, '666666')
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def generate_excel_report(
    operational_results: Dict[str, Dict[str, Any]],
    general_inputs,
    tech_inputs: Dict[str, Any] | None = None,
    cost_results: Dict[str, Dict[str, Any]] | None = None,
    cost_general=None,
    cost_inputs: Dict[str, Any] | None = None,
) -> BytesIO:
    """Genera un reporte Excel con hojas de Inputs, Operación y (opcionalmente) Costos.

    Retorna un BytesIO listo para st.download_button.
    """
    wb = Workbook()

    # ── Hoja 1: Inputs ──────────────────────────────────────────────────────
    ws_in = wb.active
    ws_in.title = 'Inputs'
    _build_inputs_sheet(ws_in, general_inputs, tech_inputs, cost_general, cost_inputs)

    # ── Hoja 2: Operación ────────────────────────────────────────────────────
    ws_op = wb.create_sheet('Operación')
    _build_operation_sheet(ws_op, operational_results)

    # ── Hoja 3: Costos (opcional) ────────────────────────────────────────────
    if cost_results:
        ws_co = wb.create_sheet('Costos')
        _build_costs_sheet(ws_co, cost_results, operational_results, cost_general)

    # Guardar en buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Constructores de hojas
# ---------------------------------------------------------------------------

def _build_inputs_sheet(ws, general_inputs, tech_inputs, cost_general, cost_inputs) -> None:
    """Construye la hoja de inputs."""
    row = 1
    n_cols = 3  # Columna de parámetro + valor (+ posible extra)

    # --- Encabezado ---
    ws.cell(row=row, column=1, value='Parámetro').font = _SECTION_FONT
    ws.cell(row=row, column=1).fill = _SECTION_FILL
    ws.cell(row=row, column=1).border = _THIN_BORDER
    ws.cell(row=row, column=2, value='Valor').font = _SECTION_FONT
    ws.cell(row=row, column=2).fill = _SECTION_FILL
    ws.cell(row=row, column=2).border = _THIN_BORDER
    ws.cell(row=row, column=2).alignment = _CENTER
    row += 1

    # --- Parámetros generales de ruta ---
    row = _write_section_header(ws, row, '🚌 Parámetros Generales de Ruta', 2)
    if general_inputs is not None:
        gi_dict = asdict(general_inputs)
        for k, v in gi_dict.items():
            label = _INPUT_LABELS.get(k, k)
            ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
            ws.cell(row=row, column=1).border = _THIN_BORDER
            ws.cell(row=row, column=2, value=_format_value(v)).font = _NORMAL_FONT
            ws.cell(row=row, column=2).border = _THIN_BORDER
            ws.cell(row=row, column=2).alignment = _CENTER
            row += 1

    # --- Parámetros por tecnología ---
    if tech_inputs:
        row += 1
        row = _write_section_header(ws, row, '⚙️ Parámetros por Tecnología', 2)
        for key, inp_obj in tech_inputs.items():
            tech_name = TECH_NAMES.get(key, key)
            icon = TECH_ICONS.get(tech_name, '🚌')
            # Sub-encabezado de tecnología
            cell = ws.cell(row=row, column=1, value=f'{icon} {tech_name}')
            cell.font = _HEADER_FONT
            cell.fill = _tech_header_fill(tech_name)
            cell.border = _THIN_BORDER
            cell2 = ws.cell(row=row, column=2)
            cell2.fill = _tech_header_fill(tech_name)
            cell2.border = _THIN_BORDER
            row += 1

            inp_dict = asdict(inp_obj)
            for k, v in inp_dict.items():
                label = _INPUT_LABELS.get(k, k)
                ws.cell(row=row, column=1, value=f'    {label}').font = _LABEL_FONT
                ws.cell(row=row, column=1).border = _THIN_BORDER
                display_v = v if v is not None else 'Auto (optimizar)'
                ws.cell(row=row, column=2, value=_format_value(display_v)).font = _NORMAL_FONT
                ws.cell(row=row, column=2).border = _THIN_BORDER
                ws.cell(row=row, column=2).alignment = _CENTER
                row += 1

    # --- Parámetros financieros ---
    if cost_general is not None:
        row += 1
        row = _write_section_header(ws, row, '💰 Parámetros Financieros Generales', 2)
        cg_dict = asdict(cost_general)
        for k, v in cg_dict.items():
            label = _INPUT_LABELS.get(k, k)
            ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
            ws.cell(row=row, column=1).border = _THIN_BORDER
            ws.cell(row=row, column=2, value=_format_value(v)).font = _NORMAL_FONT
            ws.cell(row=row, column=2).border = _THIN_BORDER
            ws.cell(row=row, column=2).alignment = _CENTER
            row += 1

    # --- Inputs de costos por tecnología ---
    if cost_inputs:
        row += 1
        row = _write_section_header(ws, row, '💶 Costos por Tecnología', 2)
        for key, parts in cost_inputs.items():
            tech_name = TECH_NAMES.get(key, key)
            icon = TECH_ICONS.get(tech_name, '🚌')
            cell = ws.cell(row=row, column=1, value=f'{icon} {tech_name}')
            cell.font = _HEADER_FONT
            cell.fill = _tech_header_fill(tech_name)
            cell.border = _THIN_BORDER
            cell2 = ws.cell(row=row, column=2)
            cell2.fill = _tech_header_fill(tech_name)
            cell2.border = _THIN_BORDER
            row += 1

            for sub_label, sub_obj in parts.items():
                # sub_label: 'capex' o 'opex'
                ws.cell(row=row, column=1, value=f'  {sub_label.upper()}').font = _SECTION_FONT
                ws.cell(row=row, column=1).fill = _SECTION_FILL
                ws.cell(row=row, column=1).border = _THIN_BORDER
                ws.cell(row=row, column=2).fill = _SECTION_FILL
                ws.cell(row=row, column=2).border = _THIN_BORDER
                row += 1
                sub_dict = asdict(sub_obj)
                for k, v in sub_dict.items():
                    label = _INPUT_LABELS.get(k, k)
                    ws.cell(row=row, column=1, value=f'    {label}').font = _LABEL_FONT
                    ws.cell(row=row, column=1).border = _THIN_BORDER
                    ws.cell(row=row, column=2, value=_format_value(v)).font = _NORMAL_FONT
                    ws.cell(row=row, column=2).border = _THIN_BORDER
                    ws.cell(row=row, column=2).alignment = _CENTER
                    row += 1

    _auto_width(ws)
    ws.freeze_panes = 'A2'


def _build_operation_sheet(ws, operational_results: Dict[str, Dict[str, Any]]) -> None:
    """Construye la hoja de resultados operativos."""
    if not operational_results:
        return

    # Recopilar todas las claves de métricas (unión de todas las tecnologías)
    all_keys: list[str] = []
    seen: set[str] = set()
    for data in operational_results.values():
        for k in data:
            if k != 'tecnologia' and k not in seen:
                all_keys.append(k)
                seen.add(k)

    tech_keys = list(operational_results.keys())
    n_techs = len(tech_keys)

    # Fila 1: encabezados
    ws.cell(row=1, column=1, value='Métrica').font = _SECTION_FONT
    ws.cell(row=1, column=1).fill = _SECTION_FILL
    ws.cell(row=1, column=1).border = _THIN_BORDER

    for j, tkey in enumerate(tech_keys):
        data = operational_results[tkey]
        tech_name = data.get('tecnologia', TECH_NAMES.get(tkey, tkey))
        icon = TECH_ICONS.get(tech_name, '🚌')
        col = j + 2
        cell = ws.cell(row=1, column=col, value=f'{icon} {tech_name}')
        cell.font = _HEADER_FONT
        cell.fill = _tech_header_fill(tech_name)
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER

    # Filas de datos
    for i, k in enumerate(all_keys):
        row = i + 2
        label = _OP_LABELS.get(k, k)
        is_total = 'total' in k.lower() or k == 'flota_requerida'

        cell_label = ws.cell(row=row, column=1, value=label)
        cell_label.font = _TOTAL_FONT if is_total else _LABEL_FONT
        cell_label.border = _THIN_BORDER
        if is_total:
            cell_label.fill = _TOTAL_FILL

        for j, tkey in enumerate(tech_keys):
            col = j + 2
            v = operational_results[tkey].get(k)
            cell = ws.cell(row=row, column=col, value=_format_value(v) if v is not None else 'N/A')
            cell.font = _TOTAL_FONT if is_total else _NORMAL_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER
            if is_total:
                cell.fill = _TOTAL_FILL
            # Formato numérico
            if isinstance(v, float):
                cell.number_format = '#,##0.00'
            elif isinstance(v, int):
                cell.number_format = '#,##0'

    _auto_width(ws)
    ws.freeze_panes = 'B2'


def _build_costs_sheet(ws, cost_results, operational_results, cost_general) -> None:
    """Construye la hoja de resultados de costos."""
    if not cost_results:
        return

    tech_keys = list(cost_results.keys())
    n_techs = len(tech_keys)
    horizonte = cost_general.horizonte_anios if cost_general else 15
    dias = cost_general.dias_operacion_anio if cost_general else 365

    # Encabezados
    ws.cell(row=1, column=1, value='Concepto').font = _SECTION_FONT
    ws.cell(row=1, column=1).fill = _SECTION_FILL
    ws.cell(row=1, column=1).border = _THIN_BORDER

    for j, tkey in enumerate(tech_keys):
        tech_name = TECH_NAMES.get(tkey, tkey)
        icon = TECH_ICONS.get(tech_name, '🚌')
        col = j + 2
        cell = ws.cell(row=1, column=col, value=f'{icon} {tech_name}')
        cell.font = _HEADER_FONT
        cell.fill = _tech_header_fill(tech_name)
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER

    row = 2

    # --- Sección CAPEX ---
    row = _write_section_header(ws, row, 'CAPEX', n_techs + 1)
    capex_keys = ['vehiculos', 'cargadores_cabecera', 'cargadores_patio',
                  'subestacion', 'estacion_h2', 'infra_deposito', 'total_capex']
    for k in capex_keys:
        label = _COST_LABELS.get(k, k)
        is_total = k == 'total_capex'
        cell_l = ws.cell(row=row, column=1, value=label)
        cell_l.font = _TOTAL_FONT if is_total else _LABEL_FONT
        cell_l.border = _THIN_BORDER
        if is_total:
            cell_l.fill = _TOTAL_FILL
        for j, tkey in enumerate(tech_keys):
            col = j + 2
            v = cost_results[tkey]['capex'].get(k, 0)
            cell = ws.cell(row=row, column=col, value=round(v, 2))
            cell.font = _TOTAL_FONT if is_total else _NORMAL_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER
            cell.number_format = '#,##0 €'
            if is_total:
                cell.fill = _TOTAL_FILL
        row += 1

    # --- Sección OPEX ---
    row = _write_section_header(ws, row, 'OPEX (anual)', n_techs + 1)
    opex_keys = ['combustible_energia', 'mantenimiento', 'bateria_anual', 'total_opex_anual']
    for k in opex_keys:
        label = _COST_LABELS.get(k, k)
        is_total = k == 'total_opex_anual'
        cell_l = ws.cell(row=row, column=1, value=label)
        cell_l.font = _TOTAL_FONT if is_total else _LABEL_FONT
        cell_l.border = _THIN_BORDER
        if is_total:
            cell_l.fill = _TOTAL_FILL
        for j, tkey in enumerate(tech_keys):
            col = j + 2
            v = cost_results[tkey]['opex'].get(k, 0)
            cell = ws.cell(row=row, column=col, value=round(v, 2))
            cell.font = _TOTAL_FONT if is_total else _NORMAL_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER
            cell.number_format = '#,##0 €'
            if is_total:
                cell.fill = _TOTAL_FILL
        row += 1

    # --- Sección TCO ---
    row = _write_section_header(ws, row, f'TCO ({horizonte} años)', n_techs + 1)

    # TCO total
    cell_l = ws.cell(row=row, column=1, value='TCO Total')
    cell_l.font = _TOTAL_FONT
    cell_l.border = _THIN_BORDER
    cell_l.fill = _TOTAL_FILL
    for j, tkey in enumerate(tech_keys):
        col = j + 2
        v = cost_results[tkey]['tco']['total_tco']
        cell = ws.cell(row=row, column=col, value=round(v, 2))
        cell.font = _TOTAL_FONT
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER
        cell.number_format = '#,##0 €'
        cell.fill = _TOTAL_FILL
    row += 1

    # €/km
    cell_l = ws.cell(row=row, column=1, value='Costo por km (€/km)')
    cell_l.font = _LABEL_FONT
    cell_l.border = _THIN_BORDER
    for j, tkey in enumerate(tech_keys):
        col = j + 2
        tco = cost_results[tkey]['tco']['total_tco']
        km_dia = operational_results[tkey]['km_totales_dia']
        km_total = km_dia * dias * horizonte
        v = tco / km_total if km_total > 0 else 0
        cell = ws.cell(row=row, column=col, value=round(v, 4))
        cell.font = _NORMAL_FONT
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER
        cell.number_format = '#,##0.00 €'
    row += 1

    # TCO / bus
    cell_l = ws.cell(row=row, column=1, value='TCO por bus (€/bus)')
    cell_l.font = _LABEL_FONT
    cell_l.border = _THIN_BORDER
    for j, tkey in enumerate(tech_keys):
        col = j + 2
        tco = cost_results[tkey]['tco']['total_tco']
        flota = operational_results[tkey]['flota_requerida']
        v = tco / flota if flota > 0 else 0
        cell = ws.cell(row=row, column=col, value=round(v, 2))
        cell.font = _NORMAL_FONT
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER
        cell.number_format = '#,##0 €'
    row += 1

    # --- Evolución anual TCO ---
    row += 1
    row = _write_section_header(ws, row, 'Evolución anual TCO', n_techs + 1)

    ws.cell(row=row, column=1, value='Año').font = _SECTION_FONT
    ws.cell(row=row, column=1).fill = _SECTION_FILL
    ws.cell(row=row, column=1).border = _THIN_BORDER
    for j, tkey in enumerate(tech_keys):
        tech_name = TECH_NAMES.get(tkey, tkey)
        col = j + 2
        cell = ws.cell(row=row, column=col, value=f'Acumulado')
        cell.font = _HEADER_FONT
        cell.fill = _tech_header_fill(tech_name)
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER
    row += 1

    for y in range(horizonte + 1):
        ws.cell(row=row, column=1, value=y).font = _NORMAL_FONT
        ws.cell(row=row, column=1).border = _THIN_BORDER
        ws.cell(row=row, column=1).alignment = _CENTER
        for j, tkey in enumerate(tech_keys):
            col = j + 2
            anual = cost_results[tkey]['tco']['anual']
            v = anual[y]['acumulado'] if y < len(anual) else 0
            cell = ws.cell(row=row, column=col, value=round(v, 2))
            cell.font = _NORMAL_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER
            cell.number_format = '#,##0 €'
        row += 1

    _auto_width(ws)
    ws.freeze_panes = 'B2'
